from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime, date, timedelta
from decimal import Decimal
import unicodedata
from typing import Any, BinaryIO

from openpyxl import load_workbook

from software_activities.domain.entities.software_activity import (
    SoftwareActivity,
    BeneficiaryBreakdown,
)
from software_activities.domain.ports.software_activity_repository import (
    SoftwareActivityRepository,
)


@dataclass(frozen=True)
class ImportResult:
    created: int
    skipped_empty_rows: int


def _excel_serial_to_date(value: Any) -> date | None:
    """
    Convierte seriales de Excel (e.g., 45776) a date.
    Excel usa 1899-12-30 como base (por el bug de 1900).
    """
    if value is None or value == "":
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    try:
        n = float(value)
    except Exception:
        return None
    base = date(1899, 12, 30)
    return base + timedelta(days=int(n))


def _to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def _to_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        raw = str(value)
        if "." in raw:
            return Decimal(raw).quantize(Decimal("1.00"))
        return Decimal(raw)
    except Exception:
        return None


def _norm(s: str) -> str:
    s = (s or "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", s).upper()


def _ffill_row(values: list[str]) -> list[str]:
    """
    Forward-fill de celdas vacías.

    Útil para merges: solo la primera celda tiene valor.
    """
    out: list[str] = []
    last = ""
    for v in values:
        if v:
            last = v
            out.append(v)
        else:
            out.append(last)
    return out


def _build_composite_headers(
    ws, header_rows: list[int], ffill_rows: set[int] | None = None
) -> list[str]:
    """
    Construye encabezados por columna concatenando varias filas (header multinivel).
    Ej: "ESTUDIANTES CUCUTA Programa X Tecnico"
    """
    max_col = ws.max_column
    rows = []
    for r in header_rows:
        vals = []
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is None:
                vals.append("")
            else:
                vals.append(str(v).strip().replace("\n", " "))
        if ffill_rows and r in ffill_rows:
            vals = _ffill_row(vals)
        rows.append(vals)

    headers: list[str] = []
    for c in range(max_col):
        parts = []
        for rvals in rows:
            p = rvals[c].strip()
            if p and p not in parts:
                parts.append(p)
        headers.append(" ".join(parts).strip())
    return headers


def _detect_header_row(ws) -> int | None:
    # busca una fila que contenga AÑO y SEMESTRE
    for r in range(1, 15):
        row = [str(c.value or "") for c in ws[r]]
        joined = _norm(" ".join(row))
        # sin tildes: ANO / SEMESTRE
        if "ANO" in joined and "SEMESTRE" in joined:
            return r
    return None


def _parse_breakdown_header(header: str) -> tuple[str, str, str, str] | None:
    """
    Intenta inferir: population, campus, program, level desde encabezados tipo:
    "ESTUDIANTES CÚCUTA PROGRAMA X TÉCNICO", etc.
    """
    h = _norm(header)
    population = None
    if "ESTUDIANTES" in h:
        population = "students"
    elif "GRADUADOS" in h:
        population = "graduates"
    elif "PROFESOR" in h and "TIPO DE BENEFICIARIO" not in h:
        # Bloque PROFESOR por programa (sin campus/nível en la plantilla)
        population = "professor"
    else:
        return None

    campus = ""
    if population in ("students", "graduates"):
        campus = (
            "CÚCUTA"
            if "CÚCUTA" in h or "CUCUTA" in h
            else ("OCAÑA" if "OCAÑA" in h or "OCANA" in h else "")
        )
        if not campus:
            return None
    else:
        campus = "N/A"

    level = ""
    if "TÉCNICO" in h or "TECNICO" in h:
        level = "técnico"
    elif "TECNÓLOGO" in h or "TECNOLOGO" in h:
        level = "tecnólogo"
    elif "PROFESIONAL" in h:
        level = "profesional"
    else:
        # Graduados/profesor en la plantilla no traen nivel: lo tratamos como total
        if population in ("graduates", "professor"):
            level = "total"
        else:
            return None

    # programa: quitamos tokens conocidos
    program = h
    for token in [
        "TIPO DE BENEFICIARIO",
        "ESTUDIANTES",
        "GRADUADOS",
        "PROFESOR",
        "CÚCUTA",
        "CUCUTA",
        "OCAÑA",
        "OCANA",
        "TÉCNICO",
        "TECNICO",
        "TECNÓLOGO",
        "TECNOLOGO",
        "PROFESIONAL",
        "PROGRAMA",
    ]:
        program = program.replace(token, " ")
    program = re.sub(r"\s+", " ", program).strip().title()
    if not program:
        program = "Sin especificar"
    return population, campus, program, level


class ImportSoftwareActivitiesFromExcelUseCase:
    """
    Importa la hoja "Software" (o la primera si no existe) y crea actividades.
    - Detecta la fila de encabezado automáticamente.
    - Lee datos desde la siguiente fila hasta el final.
    """

    def __init__(self, repository: SoftwareActivityRepository):
        self.repository = repository

    def execute(self, file_obj: BinaryIO) -> ImportResult:
        # read_only=False para poder leer merges/encabezados multinivel
        # con más fiabilidad.
        wb = load_workbook(filename=file_obj, data_only=True, read_only=False)
        ws = wb["Software"] if "Software" in wb.sheetnames else wb.worksheets[0]

        header_row = _detect_header_row(ws)
        if header_row is None:
            # fallback al formato descrito: datos desde fila 7, encabezados antes
            header_row = 5

        # En la plantilla real hay dos filas de encabezado (row 5 y 6).
        # Usamos encabezados compuestos de (row4,row5,row6) cuando existan.
        header_rows = [
            r for r in (header_row - 1, header_row, header_row + 1) if r >= 1
        ]
        # Importante: NO forward-fill en la fila de niveles (header_row+1),
        # porque las celdas vacías separan bloques (si no, se "arrastra" el nivel).
        headers = _build_composite_headers(
            ws,
            header_rows=header_rows,
            ffill_rows={header_row - 1, header_row},
        )

        # si existe sub-encabezado en header_row+1, datos comienzan en +2
        data_start_row = header_row + 1
        sub_joined = _norm(
            " ".join(
                str(ws.cell(header_row + 1, c).value or "")
                for c in range(1, ws.max_column + 1)
            )
        )
        if "ID_CINE" in sub_joined or "NUM_HORAS" in sub_joined or "TECNOLOGO" in sub_joined:
            data_start_row = header_row + 2

        header_map = {
            h.strip(): idx for idx, h in enumerate(headers) if str(h or "").strip()
        }

        def col(name: str) -> int | None:
            # match flexible por normalización
            n = _norm(name)
            for h, idx in header_map.items():
                if _norm(h) == n:
                    return idx
            # fallback: contiene (útil cuando el encabezado es multinivel)
            for h, idx in header_map.items():
                if n and n in _norm(h):
                    return idx
            return None

        # columnas base (si faltan, usamos posición por defecto)
        idx_year = col("AÑO") or 0
        idx_sem = col("SEMESTRE") or 1
        idx_start = col("FECHA INICIO DE LA ACTIVIDAD") or 2
        idx_end = col("FECHA FIN DE LA ACTIVIDAD") or 3
        idx_place = col("LUGAR DE EJECUCION DE LA ACTIVIDAD") or 4
        idx_campus = col("SEDE: CÚCUTA / OCAÑA") or 5
        idx_name = col("NOMBRE_DE LA ACTIVIDAD")
        if idx_name is None:
            idx_name = col("NOMBRE_DE LA ACTIVDAD") or 6
        idx_agreement = (
            col(
                "LA ACTIVIDAD SE DESARROLLO EN MARCO DE UN CONVENIO-  "
                "DETALLE EL NOMBRE DE LA ENTIDAD"
            )
            or 7
        )
        idx_desc = col("DESCRIPCIÓN") or 8
        idx_cine_name = (
            col(
                "CLASIFICACIÓN INTERNACIONAL NORMALIZADA DE LA EDUCACIÓN DE SUPERIOR"
            )
            or 9
        )
        # En la plantilla real, el CINE viene en la misma columna J
        # (como "613  Desarrollo...").
        idx_cine_id = col("|ID_CINE_CAMPO_DETALLADO") or 9
        idx_hours = col("NUM_HORAS") or 10
        idx_activity_type = col("ID_TIPO_ ACTIVIDAD") or 11
        idx_course_value = (
            col("VALOR_CURSO (COSTO POR PERSONA DEL EVENTO- INSCRIPCIÓN )") or 12
        )
        idx_teacher_doc_type = (
            col("ID_TIPO_DOCUMENTO DOCENTE QUE IMPARTIO  EL CURSO") or 13
        )
        idx_teacher_doc_num = col("NUM_DOCUMENTO DOCENTE QUE IMPARTIO EL CURSO") or 14

        # campos extra (si existen)
        idx_total_benef = col("TOTAL BENEFICIAIROS") or 20
        idx_prof = col("3 PROFESOR") or 17
        idx_admin = col("4 ADMINISTRATIVO IES") or 18
        idx_external = col("5PERSONA NO VINCULADA") or 19

        idx_speaker_name = col("NOMBRES Y APELLIDOS")
        idx_speaker_origin = col("PROCEDENCIA")
        idx_speaker_company = col("EMPRESA QUE REPRESENTA")

        idx_cons_entity = col("NOMBRE_ENTIDAD")
        idx_cons_sector = col("ID_SECTOR_CONSULTORIA")
        idx_cons_value = col("VALOR")

        idx_ev_plan = col("FORMATO PLANEACIÓN DE EVENTOS")
        idx_ev_att = col(
            "CONTROL ASISTENCIA ACTIVIDADES ACADEMICAS EXTRACURRICULARES"
        )
        idx_ev_prog = col(
            "FORMATO GUÍA PARA EL DISEÑO DE PROGRAMAS DE EDUCACIÓN CONTINUADA (Diplomados)"
        )
        idx_ev_av = col("REGISTRO AUDIOVISUAL")

        activities: list[SoftwareActivity] = []
        breakdowns_by_idx: dict[int, list[BeneficiaryBreakdown]] = {}
        skipped = 0

        for row in ws.iter_rows(min_row=data_start_row, values_only=True):
            year = _to_int(row[idx_year] if idx_year < len(row) else None)
            semester = _to_int(row[idx_sem] if idx_sem < len(row) else None)
            # fila vacía -> skip
            if not year and not semester and not (
                row[idx_name] if idx_name < len(row) else None
            ):
                skipped += 1
                continue

            start_date = _excel_serial_to_date(
                row[idx_start] if idx_start < len(row) else None
            )
            end_date = _excel_serial_to_date(
                row[idx_end] if idx_end < len(row) else None
            )
            campus = (
                str(row[idx_campus] or "").strip()
                if idx_campus < len(row)
                else ""
            )
            campus = campus or "CÚCUTA"

            activity = SoftwareActivity(
                id=None,
                career=None,
                year=year or 0,
                semester=semester or 0,
                start_date=start_date,
                end_date=end_date,
                execution_place=(
                    str(row[idx_place] or "").strip()
                    if idx_place < len(row)
                    else ""
                ),
                campus=campus,
                activity_name=(
                    str(row[idx_name] or "").strip() if idx_name < len(row) else ""
                ),
                agreement_entity=(
                    str(row[idx_agreement]).strip()
                    if idx_agreement < len(row) and row[idx_agreement]
                    else None
                ),
                description=(
                    str(row[idx_desc]).strip()
                    if idx_desc < len(row) and row[idx_desc]
                    else None
                ),
                cine_isced_name=(
                    str(row[idx_cine_name]).strip()
                    if idx_cine_name < len(row) and row[idx_cine_name]
                    else None
                ),
                # en la plantilla, el campo de CINE suele venir como "613  Desarrollo..."
                cine_field_detailed_id=(
                    str(row[idx_cine_id]).strip()
                    if idx_cine_id < len(row) and row[idx_cine_id]
                    else (
                        str(row[idx_cine_name]).strip().split(" ", 1)[0]
                        if idx_cine_name < len(row)
                        and row[idx_cine_name]
                        and str(row[idx_cine_name]).strip().split(" ", 1)[0].isdigit()
                        else None
                    )
                ),
                num_hours=_to_int(row[idx_hours] if idx_hours < len(row) else None),
                activity_type=(
                    str(row[idx_activity_type]).strip()
                    if idx_activity_type < len(row) and row[idx_activity_type]
                    else None
                ),
                course_value=_to_decimal(row[idx_course_value] if idx_course_value < len(row) else None),
                teacher_document_type=(
                    str(row[idx_teacher_doc_type]).strip()
                    if idx_teacher_doc_type < len(row) and row[idx_teacher_doc_type]
                    else None
                ),
                teacher_document_number=(
                    str(row[idx_teacher_doc_num]).strip()
                    if idx_teacher_doc_num < len(row) and row[idx_teacher_doc_num]
                    else None
                ),
                total_beneficiaries=_to_int(
                    row[idx_total_benef]
                    if idx_total_benef is not None and idx_total_benef < len(row)
                    else None
                ),
                professors_count=_to_int(
                    row[idx_prof]
                    if idx_prof is not None and idx_prof < len(row)
                    else None
                ),
                administrative_count=_to_int(
                    row[idx_admin]
                    if idx_admin is not None and idx_admin < len(row)
                    else None
                ),
                external_people_count=_to_int(
                    row[idx_external]
                    if idx_external is not None and idx_external < len(row)
                    else None
                ),
                speaker_full_name=(
                    str(row[idx_speaker_name]).strip()
                    if idx_speaker_name is not None
                    and idx_speaker_name < len(row)
                    and row[idx_speaker_name]
                    else None
                ),
                speaker_origin=(
                    str(row[idx_speaker_origin]).strip()
                    if idx_speaker_origin is not None
                    and idx_speaker_origin < len(row)
                    and row[idx_speaker_origin]
                    else None
                ),
                speaker_company=(
                    str(row[idx_speaker_company]).strip()
                    if idx_speaker_company is not None
                    and idx_speaker_company < len(row)
                    and row[idx_speaker_company]
                    else None
                ),
                consultancy_entity_name=(
                    str(row[idx_cons_entity]).strip()
                    if idx_cons_entity is not None
                    and idx_cons_entity < len(row)
                    and row[idx_cons_entity]
                    else None
                ),
                consultancy_sector_id=(
                    str(row[idx_cons_sector]).strip()
                    if idx_cons_sector is not None
                    and idx_cons_sector < len(row)
                    and row[idx_cons_sector]
                    else None
                ),
                consultancy_value=_to_decimal(
                    row[idx_cons_value]
                    if idx_cons_value is not None and idx_cons_value < len(row)
                    else None
                ),
                evidence_event_planning=(
                    str(row[idx_ev_plan]).strip().upper() == "SI"
                    if idx_ev_plan is not None
                    and idx_ev_plan < len(row)
                    and row[idx_ev_plan] is not None
                    else False
                ),
                evidence_attendance_control=(
                    str(row[idx_ev_att]).strip().upper() == "SI"
                    if idx_ev_att is not None
                    and idx_ev_att < len(row)
                    and row[idx_ev_att] is not None
                    else False
                ),
                evidence_program_design_guide=(
                    str(row[idx_ev_prog]).strip().upper() == "SI"
                    if idx_ev_prog is not None
                    and idx_ev_prog < len(row)
                    and row[idx_ev_prog] is not None
                    else False
                ),
                evidence_audiovisual_record=(
                    str(row[idx_ev_av]).strip().upper() == "SI"
                    if idx_ev_av is not None
                    and idx_ev_av < len(row)
                    and row[idx_ev_av] is not None
                    else False
                ),
            )
            temp_idx = len(activities)
            activities.append(activity)

            # breakdowns detectados desde headers
            breakdowns: list[BeneficiaryBreakdown] = []
            for i, h in enumerate(headers):
                parsed = _parse_breakdown_header(h)
                if not parsed:
                    continue
                if i >= len(row):
                    continue
                count = _to_int(row[i])
                if count is None:
                    continue
                population, campus_b, program, level = parsed
                breakdowns.append(
                    BeneficiaryBreakdown(
                        id=None,
                        activity_id=None,
                        population=population,
                        campus=campus_b,
                        program=program,
                        level=level,
                        count=count,
                    )
                )
            if breakdowns:
                breakdowns_by_idx[temp_idx] = breakdowns

        created = self.repository.bulk_create(
            activities=activities,
            breakdowns_by_temp_index=breakdowns_by_idx,
        )
        return ImportResult(created=created, skipped_empty_rows=skipped)
