from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from software_activities.domain.entities.software_activity import (
    BeneficiaryBreakdown,
)
from software_activities.domain.ports.software_activity_repository import (
    SoftwareActivityRepository,
)


@dataclass(frozen=True)
class ExportResult:
    filename: str
    content_type: str
    data: bytes


def _yes_no(v: bool) -> str:
    return "SI" if bool(v) else "NO"


def _fmt_date(v: date | None):
    # openpyxl soporta date y Excel lo renderiza como fecha (no serial visible)
    return v if v else None


def _collect_breakdown_columns(
    all_breakdowns: Iterable[BeneficiaryBreakdown],
) -> list[tuple[str, str, str, str]]:
    """
    Columnas dinámicas (población, campus, programa, nivel).
    Se ordena para mantener un layout estable.
    """
    uniq = {
        (b.population, b.campus, b.program, b.level)
        for b in all_breakdowns
        if b and b.program and b.level
    }
    pop_order = {"students": 0, "graduates": 1}
    campus_order = {"CÚCUTA": 0, "CUCUTA": 0, "OCAÑA": 1, "OCANA": 1}
    level_order = {"técnico": 0, "tecnólogo": 1, "profesional": 2}

    def key(t):
        pop, campus, program, level = t
        return (
            pop_order.get(pop, 99),
            campus_order.get(campus.upper(), 99),
            program.lower(),
            level_order.get(level.lower(), 99),
        )

    return sorted(list(uniq), key=key)


class ExportSoftwareActivitiesToExcelUseCase:
    """
    Exporta un Excel con el MISMO layout de la plantilla:
    - Hoja: Software
    - Encabezados: filas 1..6 (con merges)
    - Datos: desde fila 7
    """

    def __init__(self, repository: SoftwareActivityRepository):
        self.repository = repository

    def execute(self, limit: int = 5000, offset: int = 0) -> ExportResult:
        rows = self.repository.list_with_breakdowns(limit=limit, offset=offset)

        wb = Workbook()
        ws = wb.active
        ws.title = "Software"

        # ---- Layout fijo (según la plantilla real) ----
        # merges (copiados de la plantilla)
        merges = [
            "A1:CA1",
            "A5:A6",
            "B5:B6",
            "C5:C6",
            "D5:D6",
            "E5:E6",
            "F5:F6",
            "G5:G6",
            "H5:H6",
            "I5:I6",
            "P4:U4",
            "P5:P6",
            "Q5:Q6",
            "R5:R6",
            "S5:S6",
            "T5:T6",
            "U5:U6",
            "V3:BQ3",
            "V4:AR4",
            "AS4:AX4",
            "AY4:BE4",
            "BF4:BH4",
            "BI4:BO4",
            "BP4:BP6",
            "BQ4:BQ6",
            "BR3:BT4",
            "BR5:BR6",
            "BS5:BS6",
            "BT5:BT6",
            "BU3:BW4",
            "BU5:BU6",
            "BV5:BV6",
            "BW5:BW6",
            "BX3:CA4",
            "BX5:BX6",
            "BY5:BY6",
            "BZ5:BZ6",
            "CA5:CA6",
            "V5:X5",
            "Y5:Z5",
            "AA5:AC5",
            "AD5:AF5",
            "AG5:AI5",
            "AJ5:AL5",
            "AM5:AO5",
            "AP5:AR5",
            "AS5:AT5",
            "AU5:AV5",
            "AW5:AX5",
            "AY5:AY6",
            "AZ5:AZ6",
            "BA5:BA6",
            "BB5:BB6",
            "BC5:BC6",
            "BD5:BD6",
            "BE5:BE6",
            "BF5:BF6",
            "BG5:BG6",
            "BH5:BH6",
            "BI5:BI6",
            "BJ5:BJ6",
            "BK5:BK6",
            "BL5:BL6",
            "BM5:BM6",
            "BN5:BN6",
            "BO5:BO6",
        ]

        # ---- Estilos base (idénticos a plantilla) ----
        # Colores (ARGB)
        fill_title = PatternFill("solid", fgColor="FFD8D8D8")
        fill_blue = PatternFill("solid", fgColor="FFCFE2F3")  # A..I
        fill_yellow = PatternFill("solid", fgColor="FFFFF2CC")  # J5
        fill_grey_light = PatternFill("solid", fgColor="FFD8D8D8")  # bloque gris claro
        fill_grey_mid = PatternFill("solid", fgColor="FFD9D9D9")  # estudiantes cúcuta
        fill_grey_dark = PatternFill("solid", fgColor="FF595959")  # gris oscuro
        fill_peach = PatternFill("solid", fgColor="FFF7CAAC")  # tipo beneficiario
        fill_none = PatternFill()  # sin relleno (blanco)

        side_thin = Side(style="thin", color="FF000000")
        side_medium = Side(style="medium", color="FF000000")
        border_thin = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)
        border_thin_no_bottom = Border(left=side_thin, right=side_thin, top=side_thin)
        border_right_thin_bottom_thin = Border(right=side_thin, bottom=side_thin)

        font_title = Font(bold=True, size=20, name="Arial")
        # En la plantilla, los encabezados usan Arial 12.
        font_hdr_12_b = Font(bold=True, size=12, name="Arial")
        font_hdr_white_12_b = Font(bold=True, size=12, name="Arial", color="FFFFFFFF")
        font_norm_12 = Font(bold=False, size=12, name="Arial")

        align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_center_wrap_rot90 = Alignment(horizontal="center", vertical="center", wrap_text=True, textRotation=90)

        # Base: “cerrar” todas las celdas de cabecera con borde fino.
        # Importante hacerlo ANTES de merge para que las líneas no queden “a la mitad”.
        for r in range(3, 7):  # 3..6
            for c in range(1, 82):  # A..CA
                cell = ws.cell(r, c)
                cell.border = border_thin
                cell.font = font_norm_12
                cell.alignment = align_center_wrap

        # alturas de filas (como plantilla)
        ws.row_dimensions[1].height = 34.5
        ws.row_dimensions[2].height = 16.5
        ws.row_dimensions[3].height = 16.5
        ws.row_dimensions[4].height = 37.5
        ws.row_dimensions[5].height = 60.75
        ws.row_dimensions[6].height = 130.5
        ws.row_dimensions[7].height = 15.0

        # anchos de columnas (A..AX, como plantilla cuando está definido)
        col_widths_exact = {
            "A": 12.42578125,
            "B": 15.140625,
            "C": 18.85546875,
            "E": 60.0,
            "F": 18.5703125,
            "G": 50.42578125,
            "H": 28.85546875,
            "J": 55.28515625,
            "K": 19.42578125,
            "L": 26.85546875,
            "M": 21.42578125,
            "N": 30.5703125,
            "O": 24.7109375,
            "P": 5.42578125,
            "U": 8.42578125,
            "V": 4.7109375,
            "Y": 7.5703125,
            "Z": 7.28515625,
            "AA": 4.28515625,
            "AD": 4.5703125,
            "AG": 6.42578125,
            "AU": 8.42578125,
            "AV": 9.85546875,
            "AW": 6.42578125,
        }
        for letter, width in col_widths_exact.items():
            ws.column_dimensions[letter].width = width

        # Título
        ws["A1"].value = "PLANILLA REPORTE SNIES"
        ws["A1"].font = font_title
        ws["A1"].fill = fill_title
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].border = Border(left=side_medium, right=side_medium, top=side_medium, bottom=side_medium)

        # Encabezados superiores
        ws["V3"].value = "CLASIFICACIÓN DE POBLACIÓN BENEFICIADA"
        ws["BR3"].value = (
            "SI LA ACTIVIDAD ES FORMACIÓN CONTINUA, INDIQUE DATOS DEL CONFERENCISTA / PONENTE"
        )
        ws["BU3"].value = "SI LA ACTIVIDAD ES UNA CONSULTORIA DILIGENCIE:"
        ws["BX3"].value = "EVIDENCIAS DE LA ACTIVIDAD"

        ws["P4"].value = "TIPO DE BENEFICIARIO"
        ws["V4"].value = "ESTUDIANTES CÚCUTA"
        ws["AS4"].value = "ESTUDIANTES OCAÑA"
        ws["AY4"].value = "GRADUADOS CÚCUTA"
        ws["BF4"].value = "GRADUADOS OCAÑA"
        ws["BI4"].value = "PROFESOR"
        ws["BP4"].value = "ADMINISTRATIVO"
        ws["BQ4"].value = "PERSONA NO VINCULADA"

        # estilos para encabezados superiores relevantes hasta AX
        ws["V3"].fill = fill_title
        ws["V3"].font = font_hdr_12_b
        ws["V3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["V3"].border = Border(left=side_thin, right=side_thin, top=side_thin)

        ws["P4"].fill = PatternFill()  # sin relleno
        ws["P4"].font = font_norm_12
        ws["P4"].alignment = align_center_wrap
        ws["P4"].border = Border(left=side_medium, right=side_medium, top=side_medium, bottom=side_thin)

        ws["V4"].fill = fill_grey_mid
        ws["V4"].font = font_hdr_12_b
        ws["V4"].alignment = align_center_wrap
        ws["V4"].border = Border(left=side_medium, right=side_thin, top=side_medium, bottom=side_thin)

        ws["AS4"].fill = fill_grey_light
        ws["AS4"].font = font_hdr_12_b
        ws["AS4"].alignment = align_center_wrap
        ws["AS4"].border = Border(left=side_thin, top=side_medium, bottom=side_thin)

        # Fila 5 (principales)
        row5 = {
            "A5": "AÑO",
            "B5": "SEMESTRE",
            "C5": "FECHA INICIO DE LA ACTIVIDAD",
            "D5": "FECHA FIN DE LA ACTIVIDAD",
            "E5": "LUGAR DE EJECUCION DE LA ACTIVIDAD",
            "F5": "SEDE: CÚCUTA / OCAÑA",
            "G5": "NOMBRE_DE LA ACTIVDAD",
            "H5": "LA ACTIVIDAD SE DESARROLLO EN MARCO DE UN CONVENIO-  DETALLE EL NOMBRE DE LA ENTIDAD",
            "I5": "DESCRIPCIÓN",
            "J5": "CLASIFICACIÓN INTERNACIONAL NORMALIZADA DE LA EDUCACIÓN DE SUPERIOR",
            "P5": "1. ESTUDIANTE",
            "Q5": "2. GRADUADO",
            "R5": "3 PROFESOR",
            "S5": "4 ADMINISTRATIVO IES",
            "T5": "5PERSONA NO VINCULADA",
            "U5": "TOTAL BENEFICIAIROS",
            "V5": "Programa Admón Financiera",
            "Y5": "Programa Logitica Empresarial",
            "AA5": "Programa Admón Turistica y Hotelera",
            "AD5": "Programa Ing. Software",
            "AG5": "Programa Admón Negocios Internacionales (PRESENCIAL)",
            "AJ5": "Programa Admón Negocios Internacionales (DISTANCIA)",
            "AM5": "Diseño Grafico",
            "AP5": "Diseño y Admón de la moda",
            "AS5": "Programa Admón Financiera",
            "AU5": "Programa Admón Negocios Internacionales (PRESENCIAL)",
            "AW5": "Diseño Grafico",
            "AY5": "Programa Admón Financiera",
            "AZ5": "Programa Logitica Empresarial",
            "BA5": "Programa Admón Turistica y Hotelera",
            "BB5": "Programa Ing. Software",
            "BC5": "Programa Admón Negocios Internacionales",
            "BD5": "Diseño Grafico",
            "BE5": "Diseño y Admón de la moda",
            "BF5": "Programa Admón Financiera",
            "BG5": "Programa Admón Negocios Internacionales",
            "BH5": "Diseño Grafico",
            "BI5": "Programa Admón Financiera",
            "BJ5": "Programa Logitica Empresarial",
            "BK5": "Programa Admón Turistica y Hotelera",
            "BL5": "Programa Ing. Software",
            "BM5": "Programa Admón Negocios Internacionales",
            "BN5": "Diseño Grafico",
            "BO5": "Diseño y Admón de la moda",
            "BR5": "NOMBRES Y APELLIDOS",
            "BS5": "PROCEDENCIA",
            "BT5": "EMPRESA QUE REPRESENTA",
            "BU5": "NOMBRE_ENTIDAD",
            "BV5": "ID_SECTOR_CONSULTORIA",
            "BW5": "VALOR",
            "BX5": "FORMATO PLANEACIÓN DE EVENTOS",
            "BY5": "CONTROL ASISTENCIA ACTIVIDADES ACADEMICAS EXTRACURRICULARES",
            "BZ5": "FORMATO GUÍA PARA EL DISEÑO DE PROGRAMAS DE EDUCACIÓN      CONTINUADA (Diplomados)",
            "CA5": "REGISTRO AUDIOVISUAL",
        }
        for k, v in row5.items():
            ws[k].value = v

        # Fila 6 (sub-headers)
        row6 = {
            "J6": "|ID_CINE_CAMPO_DETALLADO",
            "K6": "NUM_HORAS",
            "L6": "ID_TIPO_ ACTIVIDAD",
            "M6": "VALOR_CURSO (COSTO POR PERSONA DEL EVENTO- INSCRIPCIÓN )",
            "N6": "ID_TIPO_DOCUMENTO DOCENTE QUE IMPARTIO  EL CURSO",
            "O6": "NUM_DOCUMENTO DOCENTE QUE IMPARTIO EL CURSO",
            # Estudiantes Cúcuta (V..AR)
            "V6": "Tecnico",
            "W6": "Tecnologo",
            "X6": "Profesional",
            "Y6": "Tecnico",
            "Z6": "Tecnologo",
            "AA6": "Tecnico",
            "AB6": "Tecnologo",
            "AC6": "Profesional",
            "AD6": "Tecnico",
            "AE6": "Tecnologo",
            "AF6": "Profesional",
            "AG6": "Tecnico",
            "AH6": "Tecnologo",
            "AI6": "Profesional",
            "AJ6": "Tecnico",
            "AK6": "Tecnologo",
            "AL6": "Profesional",
            "AM6": "Tecnico",
            "AN6": "Tecnologo",
            "AO6": "Profesional",
            "AP6": "Tecnico",
            "AQ6": "Tecnologo",
            "AR6": "Profesional",
            # Estudiantes Ocaña (AS..AX)
            "AS6": "Tecnologo",
            "AT6": "Profesional",
            "AU6": "Tecnologo",
            "AV6": "Profesional",
            "AW6": "Tecnologo",
            "AX6": "Profesional",
        }
        for k, v in row6.items():
            ws[k].value = v

        ws.freeze_panes = ws["A7"]

        # ---- Estilos IDENTICOS hasta la columna AX ----
        # A..I (fila 5, celdas merged hacia fila 6)
        for col in range(1, 10):  # A..I
            addr = f"{get_column_letter(col)}5"
            cell = ws[addr]
            cell.fill = fill_blue
            cell.font = font_hdr_12_b
            cell.alignment = align_left_wrap if addr in ("E5", "G5") else align_center_wrap
            cell.border = border_thin

        # J5 (CINE)
        ws["J5"].fill = fill_yellow
        ws["J5"].font = font_hdr_12_b
        ws["J5"].alignment = align_center_wrap
        ws["J5"].border = border_thin

        # J6..O6 (subheaders)
        ws["J6"].fill = fill_grey_dark
        ws["J6"].font = font_hdr_white_12_b
        ws["J6"].alignment = Alignment(vertical="center", wrap_text=True)
        ws["J6"].border = border_thin

        ws["K6"].fill = fill_grey_light
        ws["K6"].font = font_hdr_12_b
        ws["K6"].alignment = align_center_wrap
        ws["K6"].border = border_thin

        ws["L6"].fill = fill_grey_dark
        ws["L6"].font = font_hdr_white_12_b
        ws["L6"].alignment = align_center_wrap
        ws["L6"].border = border_thin

        ws["M6"].fill = fill_grey_light
        ws["M6"].font = font_hdr_12_b
        ws["M6"].alignment = align_center_wrap
        ws["M6"].border = border_thin

        ws["N6"].fill = fill_grey_dark
        ws["N6"].font = font_hdr_white_12_b
        ws["N6"].alignment = align_center_wrap
        ws["N6"].border = border_thin

        ws["O6"].fill = fill_grey_light
        ws["O6"].font = font_hdr_12_b
        ws["O6"].alignment = align_center_wrap
        ws["O6"].border = border_thin

        # Tipo de beneficiario P5..U5
        for col in range(16, 22):  # P..U
            addr = f"{get_column_letter(col)}5"
            cell = ws[addr]
            cell.fill = fill_peach
            cell.font = font_norm_12
            cell.alignment = align_center_wrap_rot90
            # bordes: P izquierda medium, U derecha medium, internos thin
            left = side_medium if addr == "P5" else side_thin
            right = side_medium if addr == "U5" else side_thin
            cell.border = Border(left=left, right=right, top=side_thin, bottom=side_thin)

        # Estudiantes Cúcuta V4..AR6 (gris medio)
        # Row 4 header
        for col in range(22, 45):  # V..AR
            addr = f"{get_column_letter(col)}4"
            c = ws[addr]
            c.fill = fill_grey_mid
            c.font = font_hdr_12_b if addr == "V4" else font_hdr_12_b
            c.alignment = align_center_wrap
            # top medium en todo el bloque, left medium solo en V
            c.border = Border(
                left=side_medium if addr == "V4" else side_thin,
                right=side_thin,
                top=side_medium,
                bottom=side_thin,
            )

        # Row 5 (programas) y Row 6 (niveles)
        for col in range(22, 45):  # V..AR
            # fila 5
            addr5 = f"{get_column_letter(col)}5"
            c5 = ws[addr5]
            c5.fill = fill_grey_mid
            c5.font = font_norm_12
            c5.alignment = align_center_wrap
            c5.border = Border(left=side_medium if addr5 == "V5" else side_thin, top=side_thin, bottom=side_thin, right=side_thin)

            # fila 6
            addr6 = f"{get_column_letter(col)}6"
            c6 = ws[addr6]
            c6.fill = fill_grey_mid
            c6.font = font_norm_12
            c6.alignment = align_center_wrap_rot90
            c6.border = Border(
                left=side_medium if addr6 == "V6" else side_thin,
                top=side_thin,
                right=side_thin,
                bottom=side_thin,
            )

        # Estudiantes Ocaña AS4..AX6 (gris claro)
        for col in range(45, 51):  # AS..AX
            # fila 4
            addr4 = f"{get_column_letter(col)}4"
            c4 = ws[addr4]
            c4.fill = fill_grey_light
            c4.font = font_hdr_12_b if addr4 == "AS4" else font_hdr_12_b
            c4.alignment = align_center_wrap
            c4.border = Border(left=side_thin, top=side_medium, bottom=side_thin, right=side_thin)

            # fila 5
            addr5 = f"{get_column_letter(col)}5"
            c5 = ws[addr5]
            c5.fill = fill_grey_light
            c5.font = font_norm_12
            c5.alignment = align_center_wrap
            c5.border = border_thin

            # fila 6
            addr6 = f"{get_column_letter(col)}6"
            c6 = ws[addr6]
            c6.fill = fill_grey_light
            c6.font = font_norm_12
            c6.alignment = align_center_wrap_rot90
            c6.border = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)

        # ---- Estilos para AY..CA (para que no queden "blancas") ----
        # Bloques superiores (fila 3): BR3..BT4, BU3..BW4, BX3..CA4
        for addr in ["BR3", "BU3", "BX3"]:
            c = ws[addr]
            c.fill = fill_grey_light
            c.font = font_hdr_12_b
            c.alignment = align_center_wrap
            c.border = border_thin

        # Fila 4: GRADUADOS CÚCUTA (AY..BE)
        for col in range(51, 57):  # AY..BE
            addr = f"{get_column_letter(col)}4"
            c = ws[addr]
            c.fill = fill_none
            c.font = font_hdr_12_b
            c.alignment = align_center_wrap
            c.border = Border(
                left=side_medium if addr == "AY4" else None,
                right=side_thin if addr == "BE4" else None,
                top=side_medium,
                bottom=side_thin,
            )

        # Fila 4: GRADUADOS OCAÑA (BF..BH)
        for col in range(58, 60 + 1):  # BF..BH
            addr = f"{get_column_letter(col)}4"
            c = ws[addr]
            c.fill = fill_peach
            c.font = font_hdr_12_b
            c.alignment = align_center_wrap
            c.border = Border(
                left=side_thin if addr == "BF4" else None,
                right=side_medium if addr == "BH4" else None,
                top=side_medium,
                bottom=side_thin,
            )

        # Fila 4: PROFESOR (BI..BO)
        for col in range(61, 67 + 1):  # BI..BO
            addr = f"{get_column_letter(col)}4"
            c = ws[addr]
            c.fill = fill_none
            c.font = font_hdr_12_b
            c.alignment = align_center_wrap
            c.border = Border(
                left=side_medium if addr == "BI4" else None,
                right=side_medium if addr == "BO4" else None,
                top=side_medium,
                bottom=side_thin,
            )

        # Fila 4: ADMINISTRATIVO / PERSONA NO VINCULADA (BP/BQ) vertical
        for addr in ["BP4", "BQ4"]:
            c = ws[addr]
            c.fill = fill_grey_light
            c.font = font_hdr_12_b
            c.alignment = align_center_wrap_rot90
            c.border = Border(left=side_medium, right=side_medium, top=side_medium)

        # Fila 5: programas/totales (rotación 90)
        # Graduados Cúcuta AY..BE (sin relleno)
        for col in range(51, 57):  # AY..BE
            addr = f"{get_column_letter(col)}5"
            c = ws[addr]
            c.fill = fill_none
            c.font = font_norm_12
            c.alignment = align_center_wrap_rot90
            c.border = Border(
                left=side_medium if addr == "AY5" else side_thin,
                right=side_thin,
                top=side_thin,
            )

        # Graduados Ocaña BF..BH (peach)
        for col in range(58, 60 + 1):  # BF..BH
            addr = f"{get_column_letter(col)}5"
            c = ws[addr]
            c.fill = fill_peach
            c.font = font_norm_12
            c.alignment = align_center_wrap_rot90
            c.border = Border(
                left=side_thin,
                right=side_medium if addr == "BH5" else side_thin,
                top=side_thin,
            )

        # Profesor BI..BO (sin relleno, extremos medium)
        for col in range(61, 67 + 1):  # BI..BO
            addr = f"{get_column_letter(col)}5"
            c = ws[addr]
            c.fill = fill_none
            c.font = font_norm_12
            c.alignment = align_center_wrap_rot90
            c.border = Border(
                left=side_medium if addr == "BI5" else side_thin,
                right=side_medium if addr == "BO5" else side_thin,
                top=side_thin,
            )

        # Fila 5: Ponente (BR..BT), Consultoría (BU..BW), Evidencias (BX..CA)
        for col in range(70, 72 + 1):  # BR..BT
            addr = f"{get_column_letter(col)}5"
            c = ws[addr]
            c.fill = fill_grey_light
            c.font = font_hdr_12_b
            c.alignment = align_center_wrap
            c.border = border_thin

        for col in range(73, 75 + 1):  # BU..BW
            addr = f"{get_column_letter(col)}5"
            c = ws[addr]
            c.fill = fill_grey_light
            c.font = font_hdr_12_b
            c.alignment = align_center_wrap
            c.border = border_thin

        for col in range(76, 79 + 1):  # BX..CA
            addr = f"{get_column_letter(col)}5"
            c = ws[addr]
            c.fill = fill_grey_light
            c.font = font_hdr_12_b
            c.alignment = align_center_wrap
            c.border = border_thin

        # Finalmente, aplicar merges (después de borders) para evitar bordes “cortados”.
        for m in merges:
            ws.merge_cells(m)

        # --- Datos ---
        start_row = 7

        # helper: map breakdowns
        def _norm_key(s: str) -> str:
            import unicodedata
            import re

            s = (s or "").strip()
            s = unicodedata.normalize("NFKD", s)
            s = "".join(ch for ch in s if not unicodedata.combining(ch))
            s = re.sub(r"\s+", " ", s).lower()
            # compat: a veces el header trae "Programa X" pero en BD guardamos solo "X"
            if s.startswith("programa "):
                s = s[len("programa ") :].strip()
            return s

        def bd_get(
            bds: list[BeneficiaryBreakdown],
            population: str,
            campus: str,
            program: str,
            level: str,
        ):
            np = _norm_key(population)
            nc = _norm_key(campus)
            nprog = _norm_key(program)
            nlvl = _norm_key(level)
            for b in bds:
                if (
                    _norm_key(b.population) == np
                    and _norm_key(b.campus) == nc
                    and _norm_key(b.program) == nprog
                    and _norm_key(b.level) == nlvl
                ):
                    return b.count
            return ""

        def bd_get_any_level(
            bds: list[BeneficiaryBreakdown],
            population: str,
            campus: str,
            program: str,
        ):
            np = _norm_key(population)
            nc = _norm_key(campus)
            nprog = _norm_key(program)
            for b in bds:
                if (
                    _norm_key(b.population) == np
                    and _norm_key(b.campus) == nc
                    and _norm_key(b.program) == nprog
                ):
                    return b.count
            return ""

        def sum_pop(bds: list[BeneficiaryBreakdown], pop: str):
            return sum(int(b.count or 0) for b in bds if b.population == pop)

        for r_idx, (a, bds) in enumerate(rows, start=start_row):
            # base
            ws[f"A{r_idx}"].value = a.year
            ws[f"B{r_idx}"].value = a.semester
            ws[f"C{r_idx}"].value = _fmt_date(a.start_date)
            ws[f"D{r_idx}"].value = _fmt_date(a.end_date)
            ws[f"E{r_idx}"].value = a.execution_place
            ws[f"F{r_idx}"].value = a.campus
            ws[f"G{r_idx}"].value = a.activity_name
            ws[f"H{r_idx}"].value = a.agreement_entity or ""
            ws[f"I{r_idx}"].value = a.description or ""
            ws[f"J{r_idx}"].value = a.cine_isced_name or ""
            ws[f"K{r_idx}"].value = a.num_hours or ""
            ws[f"L{r_idx}"].value = a.activity_type or ""
            ws[f"M{r_idx}"].value = str(a.course_value) if a.course_value is not None else ""
            ws[f"N{r_idx}"].value = a.teacher_document_type or ""
            ws[f"O{r_idx}"].value = a.teacher_document_number or ""

            # tipo beneficiario
            ws[f"P{r_idx}"].value = sum_pop(bds, "students") or ""
            ws[f"Q{r_idx}"].value = sum_pop(bds, "graduates") or ""
            ws[f"R{r_idx}"].value = a.professors_count or ""
            ws[f"S{r_idx}"].value = a.administrative_count or ""
            ws[f"T{r_idx}"].value = a.external_people_count or ""
            ws[f"U{r_idx}"].value = a.total_beneficiaries or ""

            # Estudiantes Cúcuta
            ws[f"V{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Programa Admón Financiera", "tecnico")
            ws[f"W{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Programa Admón Financiera", "tecnologo")
            ws[f"X{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Programa Admón Financiera", "profesional")
            ws[f"Y{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Programa Logitica Empresarial", "tecnico")
            ws[f"Z{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Programa Logitica Empresarial", "tecnologo")
            ws[f"AA{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Programa Admón Turistica y Hotelera", "tecnico")
            ws[f"AB{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Programa Admón Turistica y Hotelera", "tecnologo")
            ws[f"AC{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Programa Admón Turistica y Hotelera", "profesional")
            ws[f"AD{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Programa Ing. Software", "tecnico")
            ws[f"AE{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Programa Ing. Software", "tecnologo")
            ws[f"AF{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Programa Ing. Software", "profesional")
            ws[f"AG{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Programa Admón Negocios Internacionales (PRESENCIAL)", "tecnico")
            ws[f"AH{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Programa Admón Negocios Internacionales (PRESENCIAL)", "tecnologo")
            ws[f"AI{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Programa Admón Negocios Internacionales (PRESENCIAL)", "profesional")
            ws[f"AJ{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Programa Admón Negocios Internacionales (DISTANCIA)", "tecnico")
            ws[f"AK{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Programa Admón Negocios Internacionales (DISTANCIA)", "tecnologo")
            ws[f"AL{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Programa Admón Negocios Internacionales (DISTANCIA)", "profesional")
            ws[f"AM{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Diseño Grafico", "tecnico")
            ws[f"AN{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Diseño Grafico", "tecnologo")
            ws[f"AO{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Diseño Grafico", "profesional")
            ws[f"AP{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Diseño y Admón de la moda", "tecnico")
            ws[f"AQ{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Diseño y Admón de la moda", "tecnologo")
            ws[f"AR{r_idx}"].value = bd_get(bds, "students", "CÚCUTA", "Diseño y Admón de la moda", "profesional")

            # Estudiantes Ocaña (solo tecnólogo/profesional)
            ws[f"AS{r_idx}"].value = bd_get(bds, "students", "OCAÑA", "Programa Admón Financiera", "tecnologo")
            ws[f"AT{r_idx}"].value = bd_get(bds, "students", "OCAÑA", "Programa Admón Financiera", "profesional")
            ws[f"AU{r_idx}"].value = bd_get(bds, "students", "OCAÑA", "Programa Admón Negocios Internacionales (PRESENCIAL)", "tecnologo")
            ws[f"AV{r_idx}"].value = bd_get(bds, "students", "OCAÑA", "Programa Admón Negocios Internacionales (PRESENCIAL)", "profesional")
            ws[f"AW{r_idx}"].value = bd_get(bds, "students", "OCAÑA", "Diseño Grafico", "tecnologo")
            ws[f"AX{r_idx}"].value = bd_get(bds, "students", "OCAÑA", "Diseño Grafico", "profesional")

            # Graduados Cúcuta (totales por programa)
            ws[f"AY{r_idx}"].value = bd_get_any_level(bds, "graduates", "CÚCUTA", "Programa Admón Financiera")
            ws[f"AZ{r_idx}"].value = bd_get_any_level(bds, "graduates", "CÚCUTA", "Programa Logitica Empresarial")
            ws[f"BA{r_idx}"].value = bd_get_any_level(bds, "graduates", "CÚCUTA", "Programa Admón Turistica y Hotelera")
            ws[f"BB{r_idx}"].value = bd_get_any_level(bds, "graduates", "CÚCUTA", "Programa Ing. Software")
            ws[f"BC{r_idx}"].value = bd_get_any_level(bds, "graduates", "CÚCUTA", "Programa Admón Negocios Internacionales")
            ws[f"BD{r_idx}"].value = bd_get_any_level(bds, "graduates", "CÚCUTA", "Diseño Grafico")
            ws[f"BE{r_idx}"].value = bd_get_any_level(bds, "graduates", "CÚCUTA", "Diseño y Admón de la moda")

            # Graduados Ocaña
            ws[f"BF{r_idx}"].value = bd_get_any_level(bds, "graduates", "OCAÑA", "Programa Admón Financiera")
            ws[f"BG{r_idx}"].value = bd_get_any_level(bds, "graduates", "OCAÑA", "Programa Admón Negocios Internacionales")
            ws[f"BH{r_idx}"].value = bd_get_any_level(bds, "graduates", "OCAÑA", "Diseño Grafico")

            # Profesor (por programa, totales)
            ws[f"BI{r_idx}"].value = bd_get_any_level(bds, "professor", "N/A", "Programa Admón Financiera")
            ws[f"BJ{r_idx}"].value = bd_get_any_level(bds, "professor", "N/A", "Programa Logitica Empresarial")
            ws[f"BK{r_idx}"].value = bd_get_any_level(bds, "professor", "N/A", "Programa Admón Turistica y Hotelera")
            ws[f"BL{r_idx}"].value = bd_get_any_level(bds, "professor", "N/A", "Programa Ing. Software")
            ws[f"BM{r_idx}"].value = bd_get_any_level(bds, "professor", "N/A", "Programa Admón Negocios Internacionales")
            ws[f"BN{r_idx}"].value = bd_get_any_level(bds, "professor", "N/A", "Diseño Grafico")
            ws[f"BO{r_idx}"].value = bd_get_any_level(bds, "professor", "N/A", "Diseño y Admón de la moda")

            # Duplicados en la plantilla (clasificación): repetimos totales
            ws[f"BP{r_idx}"].value = a.administrative_count or ""
            ws[f"BQ{r_idx}"].value = a.external_people_count or ""

            # Ponente / formación continua
            ws[f"BR{r_idx}"].value = a.speaker_full_name or ""
            ws[f"BS{r_idx}"].value = a.speaker_origin or ""
            ws[f"BT{r_idx}"].value = a.speaker_company or ""

            # Consultoría
            ws[f"BU{r_idx}"].value = a.consultancy_entity_name or ""
            ws[f"BV{r_idx}"].value = a.consultancy_sector_id or ""
            ws[f"BW{r_idx}"].value = str(a.consultancy_value) if a.consultancy_value is not None else ""

            # Evidencias
            ws[f"BX{r_idx}"].value = _yes_no(a.evidence_event_planning)
            ws[f"BY{r_idx}"].value = _yes_no(a.evidence_attendance_control)
            ws[f"BZ{r_idx}"].value = _yes_no(a.evidence_program_design_guide)
            ws[f"CA{r_idx}"].value = _yes_no(a.evidence_audiovisual_record)

            # Formato fechas
            ws[f"C{r_idx}"].number_format = "dd/mm/yyyy"
            ws[f"D{r_idx}"].number_format = "dd/mm/yyyy"

            # Bordes (para que las líneas se vean completas en el export).
            # La plantilla aplica borde fino en la zona P..CA (y en general en la fila de datos).
            for c in range(1, 81 + 1):  # A..CA
                ws.cell(r_idx, c).border = border_thin

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return ExportResult(
            filename="planilla_reporte_snies_software.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            data=bio.read(),
        )
